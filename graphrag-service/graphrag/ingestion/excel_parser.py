"""Parse Excel tailings reports into LossForm nodes and bucket chunks."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from graphrag.constants import METAL_CU, METAL_NI
from graphrag.ingestion.domain import (
    FORM_SLUG,
    MINERAL_FORM_LABELS,
    intervention_edges_for_bucket,
    is_recoverable_form,
    normalize_size_class,
)
from graphrag.models import Chunk, GraphEdge, GraphNode
from graphrag.schema import NodeType

_SIZE_HEADER_RE = re.compile(r"^\+?\d|^-")
_INVALID_VALUES = ("#REF!", None)


@dataclass
class ExcelParseResult:
    factory: str
    source_file: str
    nodes: list[GraphNode]
    edges: list[GraphEdge]
    chunks: list[Chunk]


def _slugify(*parts: str) -> str:
    raw = "_".join(parts).lower()
    cleaned = re.sub(r"[^\w+]+", "_", raw, flags=re.UNICODE)

    return re.sub(r"_+", "_", cleaned).strip("_")


def _safe_float(value: Any) -> float | None:
    if value in _INVALID_VALUES:
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_size_header(row: tuple[Any, ...]) -> bool:
    label = row[1]

    if not isinstance(label, str):
        return False

    if "Доля потерь" not in str(row[3] or ""):
        return False

    return bool(_SIZE_HEADER_RE.match(label.strip()))


def _parse_sheet(ws, factory: str, source_file: str) -> ExcelParseResult:
    nodes: list[GraphNode] = []
    edges: list[GraphEdge] = []
    chunks: list[Chunk] = []
    current_size: str | None = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _is_size_header(row):
            current_size = normalize_size_class(str(row[1]))
            continue

        form_label = row[1]

        if current_size is None:
            continue

        if form_label not in MINERAL_FORM_LABELS:
            continue

        form_slug = FORM_SLUG[form_label]
        ni_tonnes = _safe_float(row[4])
        cu_tonnes = _safe_float(row[6])

        for metal, tonnes in ((METAL_NI, ni_tonnes), (METAL_CU, cu_tonnes)):
            if tonnes is None or tonnes <= 0:
                continue

            recoverable = is_recoverable_form(form_slug, metal)
            node_id = _slugify("lossform", factory, current_size, form_slug, metal)

            nodes.append(
                GraphNode(
                    node_id=node_id,
                    node_type=NodeType.LOSS_FORM,
                    label=f"{factory} {current_size} {form_label.strip()} {metal}",
                    attributes={
                        "factory": factory,
                        "size_class": current_size,
                        "mineral_form": form_slug,
                        "metal": metal,
                        "tonnes": round(tonnes, 2),
                        "recoverable": recoverable,
                        "source_file": source_file,
                    },
                )
            )

            chunk_text = (
                f"Фабрика {factory}. Класс крупности {current_size}. "
                f"Форма: {form_label.strip()}. Потери {metal}: {tonnes:.1f} т. "
                f"Извлекаемо: {'да' if recoverable else 'нет'}."
            )
            chunks.append(
                Chunk(
                    chunk_id=f"excel_{node_id}",
                    text=chunk_text,
                    summary=f"{factory} bucket {current_size} {form_slug} {metal}",
                    source=source_file,
                    factory=factory,
                    chunk_type="excel_bucket",
                    graph_node_ids=_chunk_graph_nodes(
                        node_id=node_id,
                        form_slug=form_slug,
                        size_class=current_size,
                        metal=metal,
                        recoverable=recoverable,
                    ),
                    metadata={
                        "sheet": ws.title,
                        "excel_row": row_idx,
                        "excel_cell_ni": f"E{row_idx}",
                        "excel_cell_cu": f"G{row_idx}",
                        "metal": metal,
                    },
                )
            )

            for source, relation, target in intervention_edges_for_bucket(
                bucket_node_id=node_id,
                form_slug=form_slug,
                size_class=current_size,
                metal=metal,
                recoverable=recoverable,
            ):
                edges.append(GraphEdge(source, target, relation))

    return ExcelParseResult(
        factory=factory,
        source_file=source_file,
        nodes=nodes,
        edges=edges,
        chunks=chunks,
    )


def _chunk_graph_nodes(
    *,
    node_id: str,
    form_slug: str,
    size_class: str,
    metal: str,
    recoverable: bool,
) -> list[str]:
    """Link excel chunk to bucket + intervention path nodes for graph retrieval."""
    node_ids = [node_id]

    if not recoverable:
        return node_ids

    for source, _relation, target in intervention_edges_for_bucket(
        bucket_node_id=node_id,
        form_slug=form_slug,
        size_class=size_class,
        metal=metal,
        recoverable=recoverable,
    ):
        if source == node_id and target not in node_ids:
            node_ids.append(target)

    return node_ids


def parse_excel_file(path: Path, factory: str) -> ExcelParseResult:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    return _parse_sheet(worksheet, factory, path.name)
