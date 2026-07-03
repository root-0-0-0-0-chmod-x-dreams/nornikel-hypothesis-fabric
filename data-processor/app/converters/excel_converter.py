from __future__ import annotations

import logging
import shutil
from pathlib import Path
from typing import Any, Dict, List

from app.models import ConvertResult
from app.utils import FileType

logger = logging.getLogger("data_processor")


def convert_xlsx(file_path: Path, original_filename: str) -> ConvertResult:
    errors: List[str] = []
    warnings: List[str] = []

    try:
        import pandas as pd
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), data_only=True)
        sheets_md: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = [[_safe_cell_value(cell.value) for cell in row] for row in ws.iter_rows()]

            if not data:
                continue

            max_cols = max((len(row) for row in data), default=0)
            normalized = []
            for row in data:
                if len(row) < max_cols:
                    row = row + [""] * (max_cols - len(row))
                normalized.append(row)

            df = pd.DataFrame(normalized)

            sheets_md.append(f"## {sheet_name}\n")
            sheets_md.append(df.to_markdown(index=False))
            sheets_md.append("")

        wb.close()
        return ConvertResult(
            success=True,
            original_filename=original_filename,
            file_type=FileType.XLSX.value,
            markdown_content="\n".join(sheets_md),
            metadata={"sheets": wb.sheetnames},
            warnings=warnings,
        )

    except ImportError:
        return ConvertResult(
            success=False,
            original_filename=original_filename,
            file_type=FileType.XLSX.value,
            errors=["openpyxl or pandas not installed. Install: pip install openpyxl pandas tabulate"],
        )
    except Exception as e:
        logger.exception("xlsx_conversion_failed")
        errors.append(str(e))
        return ConvertResult(
            success=False,
            original_filename=original_filename,
            file_type=FileType.XLSX.value,
            errors=errors,
        )


def convert_xls(file_path: Path, original_filename: str) -> ConvertResult:
    errors: List[str] = []
    warnings: List[str] = []

    try:
        xlsx_path = file_path.with_suffix(".xlsx")
        try:
            from xls2xlsx import XLS2XLSX
            XLS2XLSX(str(file_path)).to_xlsx(str(xlsx_path))
        except ImportError:
            return ConvertResult(
                success=False,
                original_filename=original_filename,
                file_type=FileType.XLS.value,
                errors=["xls2xlsx not installed. Install: pip install xls2xlsx"],
            )
        except Exception as e:
            errors.append(f"XLS→XLSX conversion failed: {e}")

        if not xlsx_path.exists():
            return ConvertResult(
                success=False,
                original_filename=original_filename,
                file_type=FileType.XLS.value,
                errors=["Failed to create XLSX from XLS"] + errors,
            )

        result = convert_xlsx(xlsx_path, original_filename)
        result.file_type = FileType.XLS.value
        result.warnings = warnings + result.warnings

        xlsx_path.unlink(missing_ok=True)
        return result

    except Exception as e:
        logger.exception("xls_conversion_failed")
        errors.append(str(e))
        return ConvertResult(
            success=False,
            original_filename=original_filename,
            file_type=FileType.XLS.value,
            errors=errors,
        )


def _safe_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value == int(value):
            return str(int(value))
    return str(value)
