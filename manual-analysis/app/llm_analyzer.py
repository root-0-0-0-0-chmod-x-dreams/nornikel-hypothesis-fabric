from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import List, Optional

import httpx

from .config import get_config
from .models import AnalysisRow

logger = logging.getLogger("manual_analysis")

_PROMPT = """Ты — эксперт по обогащению полезных ископаемых. Проанализируй таблицу данных о хвостах флотации и заполни шаблон анализа потерь.

Данные из Excel (упрощённое представление):
{excel_data}

Шаблон для заполнения (CSV-формат):
Анализ,Статья_Потерь,Металл_1_т,Металл_2_т,Доп_Инфо

Категории Анализа:
- Минералогия: потери, связанные с минералогическими формами (валлериит, силикаты, примеси в пирротине, пирит, сростки)
- Фракции: потери из-за крупности частиц (шламы/переизмельчение, крупные сростки/недоизмельчение)
- Автоматизация: потери из-за нестабильности процесса (колебания плотности, гранулометрии)

Правила:
1. Извлеки из данных фактические значения потерь в тоннах для каждого металла (колонки "Извлечение ..., т")
2. Суммируй близкие статьи потерь по категориям
3. Для Автоматизации, если нет явных данных, оцени как ~5% от общих флотационных потерь
4. В Доп_Инфо укажи краткую характеристику причины потерь

Верни ТОЛЬКО валидный JSON-массив объектов с полями:
[{{"Анализ": "...", "Статья_Потерь": "...", "Металл_1_т": число, "Металл_2_т": число, "Доп_Инфо": "..."}}]
Без текста до и после, без ```json, только массив.
"""


async def check_llm_available() -> tuple[bool, str]:
    config = get_config()
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            resp = await client.get(f"{config.llm_service_url}/health")
            if resp.status_code == 200:
                return True, ""
            return False, f"HTTP {resp.status_code}"
    except Exception as e:
        return False, str(e)


async def analyze_with_llm(file_path: Path) -> List[AnalysisRow]:
    config = get_config()

    excel_data = _prepare_excel_data(file_path)
    if not excel_data:
        raise ValueError("Failed to read Excel data")

    prompt = _PROMPT.format(excel_data=excel_data[:8000])

    async with httpx.AsyncClient(timeout=config.llm_timeout) as client:
        resp = await client.post(
            f"{config.llm_service_url}/chat",
            json={
                "model": config.llm_model,
                "messages": [
                    {"role": "system", "content": "Ты — эксперт-аналитик. Отвечай строго JSON-массивом без лишнего текста."},
                    {"role": "user", "content": prompt},
                ],
                "temperature": 0.3,
                "max_tokens": 4000,
            },
        )

    if resp.status_code != 200:
        raise RuntimeError(f"LLM service returned {resp.status_code}: {resp.text[:500]}")

    data = resp.json()
    text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
    text = text.strip()

    if text.startswith("```"):
        text = text.split("\n", 1)[-1]
        if text.endswith("```"):
            text = text[:-3]

    try:
        raw_rows = json.loads(text)
    except json.JSONDecodeError:
        import re
        match = re.search(r"\[.*\]", text, re.DOTALL)
        if match:
            raw_rows = json.loads(match.group())
        else:
            raise ValueError(f"LLM returned non-JSON: {text[:300]}")

    rows = []
    for item in raw_rows:
        try:
            rows.append(AnalysisRow(
                Анализ=str(item.get("Анализ", "")),
                Статья_Потерь=str(item.get("Статья_Потерь", "")),
                Металл_1_т=float(item.get("Металл_1_т", 0)),
                Металл_2_т=float(item.get("Металл_2_т", 0)),
                Доп_Инфо=str(item.get("Доп_Инфо", "")),
            ))
        except (ValueError, TypeError) as e:
            logger.warning("skip_llm_row", extra={"item": str(item)[:100], "error": str(e)})

    return rows


def _prepare_excel_data(file_path: Path) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        lines = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(c) if c is not None else "" for c in row]
            if any(v.strip() for v in vals):
                lines.append(" | ".join(vals))

        wb.close()
        return "\n".join(lines[:200])
    except Exception as e:
        logger.exception("excel_read_failed")
        return ""
