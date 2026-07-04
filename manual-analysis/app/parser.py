from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook

from .models import AnalysisRow

logger = logging.getLogger("manual_analysis")


def parse_excel(file_path: Path) -> List[AnalysisRow]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    rows_raw = []
    for row in ws.iter_rows(values_only=True):
        rows_raw.append(tuple(row))
    wb.close()

    label_col, m1_t, m2_t = _find_tonnage_columns(rows_raw)
    labels = _extract_labels(rows_raw, label_col)

    mineral_rows = _collect_mineral_rows(labels, m1_t, m2_t)
    fine_m1, fine_m2, coarse_m1, coarse_m2 = _collect_fraction_totals(labels, m1_t, m2_t)

    result: List[AnalysisRow] = []

    if mineral_rows:
        sil_m1 = mineral_rows.get("Силикатная форма/Валлериит", 0.0)
        pirr_m1 = mineral_rows.get("Примесь в пирротине", 0.0)
        pir_m1 = mineral_rows.get("Пирит/Другие", 0.0)
        sil_m2 = mineral_rows.get("Силикатная форма/Валлериит_m2", 0.0)
        pirr_m2 = mineral_rows.get("Примесь в пирротине_m2", 0.0)
        pir_m2 = mineral_rows.get("Пирит/Другие_m2", 0.0)

        if sil_m1 > 0 or sil_m2 > 0:
            result.append(AnalysisRow(
                Анализ="Минералогия",
                Статья_Потерь="Силикатная форма/Валлериит",
                Металл_1_т=round(sil_m1, 2),
                Металл_2_т=round(sil_m2, 2),
                Доп_Инфо="Труднофлотируемые формы",
            ))
        if pirr_m1 > 0 or pirr_m2 > 0:
            result.append(AnalysisRow(
                Анализ="Минералогия",
                Статья_Потерь="Примесь в пирротине",
                Металл_1_т=round(pirr_m1, 2),
                Металл_2_т=round(pirr_m2, 2),
                Доп_Инфо="Магнитный сульфид",
            ))
        if pir_m1 > 0 or pir_m2 > 0:
            result.append(AnalysisRow(
                Анализ="Минералогия",
                Статья_Потерь="Пирит и другие сульфиды",
                Металл_1_т=round(pir_m1, 2),
                Металл_2_т=round(pir_m2, 2),
                Доп_Инфо="Высокая плотность",
            ))

    if fine_m1 > 0 or fine_m2 > 0:
        result.append(AnalysisRow(
            Анализ="Фракции",
            Статья_Потерь="Шламы (-10 мкм) Раскрытый сульфид",
            Металл_1_т=round(fine_m1, 2),
            Металл_2_т=round(fine_m2, 2),
            Доп_Инфо="Переизмельчение, тонкие классы",
        ))
    if coarse_m1 > 0 or coarse_m2 > 0:
        result.append(AnalysisRow(
            Анализ="Фракции",
            Статья_Потерь="Крупные сростки (+74 мкм) Нераскрытый",
            Металл_1_т=round(coarse_m1, 2),
            Металл_2_т=round(coarse_m2, 2),
            Доп_Инфо="Недоизмельчение, крупные классы",
        ))

    total_loss_m1 = sum(r.metal_1_t for r in result)
    auto_m1 = max(round(total_loss_m1 * 0.05, 1), 50.0)
    total_loss_m2 = sum(r.metal_2_t for r in result)
    auto_m2 = max(round(total_loss_m2 * 0.05, 1), 30.0)

    result.append(AnalysisRow(
        Анализ="Автоматизация",
        Статья_Потерь="Колебания плотности / Гранулометрии",
        Металл_1_т=auto_m1,
        Металл_2_т=auto_m2,
        Доп_Инфо="Нестабильность питания контура",
    ))

    return result


def _find_tonnage_columns(rows_raw) -> Tuple[int, int, int]:
    label_col = 1
    m1_t = 4
    m2_t = 6

    for row in rows_raw[:5]:
        for i, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip().lower()
            if ("28" in s or "никел" in s) and ("т" in s or "извлечен" in s or "извл" in s):
                m1_t = i
            if ("29" in s or "мед" in s) and ("т" in s or "извлечен" in s or "извл" in s):
                m2_t = i

    return label_col, m1_t, m2_t


def _extract_labels(rows_raw, label_col) -> List[Tuple[str, tuple]]:
    out = []
    for row in rows_raw:
        if label_col >= len(row):
            continue
        label = row[label_col]
        if label is not None and str(label).strip():
            out.append((str(label).strip(), row))
    return out


def _val(row, col) -> float:
    if col >= len(row):
        return 0.0
    v = row[col]
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _collect_mineral_rows(labels, m1_t, m2_t) -> dict:
    totals = {}

    for label, row in labels:
        s = label.strip()

        if "Силикатная форма" in s and "Валлер" in s:
            totals["Силикатная форма/Валлериит"] = totals.get("Силикатная форма/Валлериит", 0.0) + _val(row, m1_t)
            totals["Силикатная форма/Валлериит_m2"] = totals.get("Силикатная форма/Валлериит_m2", 0.0) + _val(row, m2_t)

        elif "Примесь в пирротине" in s:
            totals["Примесь в пирротине"] = totals.get("Примесь в пирротине", 0.0) + _val(row, m1_t)
            totals["Примесь в пирротине_m2"] = totals.get("Примесь в пирротине_m2", 0.0) + _val(row, m2_t)

        elif "Пирит" in s and "сульфид" in s.lower():
            totals["Пирит/Другие"] = totals.get("Пирит/Другие", 0.0) + _val(row, m1_t)
            totals["Пирит/Другие_m2"] = totals.get("Пирит/Другие_m2", 0.0) + _val(row, m2_t)

    return totals


def _collect_fraction_totals(labels, m1_t, m2_t) -> Tuple[float, float, float, float]:
    fine_m1, fine_m2 = 0.0, 0.0
    coarse_m1, coarse_m2 = 0.0, 0.0

    current_fraction = None

    for label, row in labels:
        s = label.strip()

        if s in ("+125", "+71", "-71 + 45", "-45 + 20", "-20 + 10", " -20 + 10", "-10"):
            current_fraction = s.strip()

        if "Итого" in s and current_fraction:
            m1 = _val(row, m1_t)
            m2 = _val(row, m2_t)

            frac = current_fraction.strip()
            if frac in ("-10", "-20 + 10", " -20 + 10"):
                fine_m1 += m1
                fine_m2 += m2
            elif frac in ("+125", "+71", "-71 + 45"):
                coarse_m1 += m1
                coarse_m2 += m2
            elif frac == "-45 + 20":
                coarse_m1 += m1 * 0.5
                coarse_m2 += m2 * 0.5
                fine_m1 += m1 * 0.5
                fine_m2 += m2 * 0.5

            current_fraction = None

    return fine_m1, fine_m2, coarse_m1, coarse_m2
